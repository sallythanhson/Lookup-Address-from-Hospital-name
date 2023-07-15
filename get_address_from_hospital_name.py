import openpyxl
import requests
import json

api_key = ''


# Specify the Google Map API key and base URL for the Google Maps API above
url = 'https://maps.googleapis.com/maps/api/place/findplacefromtext/json'

# Open the Excel file and select the active worksheet
workbook = openpyxl.load_workbook('hospital_list.xlsx')
worksheet = workbook.active

# Iterate over each row in column A of the worksheet
for row in worksheet.iter_rows(min_row=2, min_col=1):
    hospital_name_cell = row[0]
    
    # Create the payload for the Google Maps API request
    params = {
        'key': api_key,
        'input': hospital_name_cell.value,
        'inputtype': 'textquery',
        'fields': 'formatted_address'
    }
    
     # Send the request to the Google Maps API and parse the response
    response = requests.get(url, params=params)
    data = json.loads(response.content)
    
    # Extract the hospital address from the API response
    if data['status'] == 'OK':
        hospital_address = data['candidates'][0]['formatted_address']
        # Write the hospital address to column B of the worksheet
        worksheet.cell(row=hospital_name_cell.row, column=2, value=hospital_address)
        print("Check address", row)
    else:
        worksheet.cell(row=hospital_name_cell.row, column=2, value='Address not found')

        
# Save the changes to the Excel file
workbook.save('hospital_list_with_addresses.xlsx')
